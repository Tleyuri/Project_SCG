"""สร้างไฟล์ Excel BOQ ตาม template `ถอดวัสดุ_3พืช.xlsx` (ข้อ 5)

1 พืช = 1 sheet, คอลัมน์: ลำดับที่ | Mat.code | ชื่อรายการ | จำนวน | หน่วยนับ | หมายเหตุ | ราคา/ชั้น | รวม
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = ["ลำดับที่", "Mat.code", "ชื่อรายการ", "จำนวน", "หน่วยนับ", "หมายเหตุ", "ราคา/ชั้น", "รวม"]


def build_workbook(boq_plants: dict, garden_location: str = "", garden_phone: str = "") -> bytes:
    """สร้าง workbook โดยแต่ละ key ใน boq_plants คือชื่อ sheet (พืช)

    boq_plants: {"ทุเรียน": {"rows": [...]}, "กาแฟ": {"rows": [...]}}
    """
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    center = Alignment(horizontal="center", vertical="center")

    for plant_name, data in boq_plants.items():
        ws = wb.create_sheet(title=plant_name[:31])

        ws.merge_cells("A1:H1")
        ws["A1"] = "รายการอุปกรณ์สวน"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = center

        ws.merge_cells("A2:H2")
        ws["A2"] = f"สวนอยู่ที่ จ.{garden_location}  Tel : {garden_phone}"
        ws["A2"].alignment = center

        header_row = 4
        for col, title in enumerate(HEADERS, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        rows = data.get("rows", [])
        for i, row in enumerate(rows, start=1):
            r = header_row + i
            ws.cell(row=r, column=1, value=i)
            ws.cell(row=r, column=2, value=row.get("mat_code", ""))
            ws.cell(row=r, column=3, value=row.get("name", ""))
            ws.cell(row=r, column=4, value=row.get("qty", 0))
            ws.cell(row=r, column=5, value=row.get("unit", ""))
            ws.cell(row=r, column=6, value=row.get("note", ""))
            ws.cell(row=r, column=7, value=row.get("unit_price", 0))
            ws.cell(row=r, column=8, value=f"=D{r}*G{r}")

        # ความกว้างคอลัมน์
        widths = [8, 12, 45, 10, 10, 35, 12, 14]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

    if not wb.sheetnames:
        wb.create_sheet(title="BOQ")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
